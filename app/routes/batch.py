import asyncio
import json
import os
import re
import uuid as _uuid
from datetime import datetime
from typing import Dict

from fastapi import APIRouter, HTTPException, Depends, UploadFile, File

from app.core.security import verify_token
from app.schemas.requests import BatchRunRequest
from app.services.agent import create_batch_agent
from app.services.prompt import get_user_system_prompt, get_prompt_version
from app.deps import get_current_user
from app.storage import get_storage_service

router = APIRouter(prefix="/api/batch", tags=["batch"])

_batch_tasks: Dict[str, dict] = {}


def _col_letter_to_index(letter: str) -> int:
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


@router.post("/upload")
async def api_batch_upload(file: UploadFile = File(...), user=Depends(get_current_user)):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")
    user_id = user["user_id"]
    safe_name = re.sub(r'[^\w\-.]', '_', file.filename)
    content = await file.read()
    storage = get_storage_service()
    rel_path = f"/.batch_uploads/{safe_name}"
    storage.write_bytes(user_id, rel_path, content)

    import openpyxl
    import io
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets_info = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        headers = [str(cell.value) if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1), [])]
        sheets_info.append({"name": sn, "headers": headers, "row_count": ws.max_row or 0})
    wb.close()
    return {"filename": safe_name, "sheets": sheets_info}


@router.post("/run")
async def api_batch_run(req: BatchRunRequest, user=Depends(get_current_user)):
    user_id = user["user_id"]
    storage = get_storage_service()
    upload_path = f"/.batch_uploads/{req.filename}"
    if not storage.exists(user_id, upload_path):
        raise HTTPException(status_code=404, detail="Excel 文件不存在，请先上传")

    if req.prompt_version_id:
        version = get_prompt_version(user_id, req.prompt_version_id)
        if not version:
            raise HTTPException(status_code=404, detail="Prompt 版本不存在")
        prompt_content = version["content"]
    else:
        prompt_content = get_user_system_prompt(user_id)

    task_id = _uuid.uuid4().hex[:10]
    result_rel = f"/.batch_results/{task_id}/result_{req.filename}"

    source_bytes = storage.read_bytes(user_id, upload_path)
    storage.write_bytes(user_id, result_rel, source_bytes)

    has_running = any(t["user_id"] == user_id and t["status"] == "running" for t in _batch_tasks.values())
    initial_status = "queued" if has_running else "running"

    task = {
        "id": task_id, "user_id": user_id, "status": initial_status,
        "total": max(0, req.end_row - req.start_row + 1), "completed": 0,
        "current_query": "", "results": [],
        "config": {
            "model": req.model, "query_col": req.query_col, "content_col": req.content_col,
            "tool_col": req.tool_col, "start_row": req.start_row, "end_row": req.end_row,
            "sheet_name": req.sheet_name,
        },
        "result_path": result_rel,
        "created_at": datetime.now().isoformat(), "finished_at": None,
        "error": None, "_cancelled": False, "_prompt_content": prompt_content,
    }
    _batch_tasks[task_id] = task

    if initial_status == "running":
        asyncio.create_task(_run_batch_task(task_id, user_id, req.model, prompt_content))

    return {"task_id": task_id, "status": initial_status, "total": task["total"]}


@router.get("/tasks")
async def api_list_batch_tasks(user=Depends(get_current_user)):
    user_id = user["user_id"]
    return [
        {"id": t["id"], "status": t["status"], "total": t["total"], "completed": t["completed"],
         "created_at": t["created_at"], "finished_at": t.get("finished_at"), "error": t.get("error")}
        for t in _batch_tasks.values() if t["user_id"] == user_id
    ]


@router.get("/tasks/{task_id}")
async def api_get_batch_task(task_id: str, user=Depends(get_current_user)):
    task = _batch_tasks.get(task_id)
    if not task or task["user_id"] != user["user_id"]:
        raise HTTPException(status_code=404, detail="任务不存在")
    return {
        "id": task["id"], "status": task["status"], "total": task["total"],
        "completed": task["completed"], "current_query": task.get("current_query", ""),
        "results": task["results"], "config": task["config"],
        "created_at": task["created_at"], "finished_at": task.get("finished_at"), "error": task.get("error"),
    }


@router.post("/tasks/{task_id}/cancel")
async def api_cancel_batch_task(task_id: str, user=Depends(get_current_user)):
    task = _batch_tasks.get(task_id)
    if not task or task["user_id"] != user["user_id"]:
        raise HTTPException(status_code=404, detail="任务不存在")
    task["_cancelled"] = True
    if task["status"] == "queued":
        task["status"] = "cancelled"
        task["finished_at"] = datetime.now().isoformat()
    return {"success": True}


@router.get("/tasks/{task_id}/download")
async def api_download_batch_result(task_id: str, token: str):
    user = verify_token(token)
    if not user:
        raise HTTPException(status_code=401, detail="无效的 token")
    task = _batch_tasks.get(task_id)
    if not task or task["user_id"] != user["user_id"]:
        raise HTTPException(status_code=404, detail="任务不存在")

    storage = get_storage_service()
    result_rel = task["result_path"]
    if not storage.exists(user["user_id"], result_rel):
        raise HTTPException(status_code=404, detail="结果文件不存在")
    return storage.file_response(
        user["user_id"], result_rel,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(result_rel), inline=False,
    )


async def _run_batch_task(task_id: str, user_id: str, model: str, prompt_content: str):
    task = _batch_tasks[task_id]
    cfg = task["config"]
    query_col_idx = _col_letter_to_index(cfg["query_col"])
    content_col_idx = _col_letter_to_index(cfg["content_col"])
    tool_col_idx = _col_letter_to_index(cfg["tool_col"])

    storage = get_storage_service()
    result_rel = task["result_path"]

    try:
        agent = create_batch_agent(user_id, model, prompt_content)

        result_bytes = storage.read_bytes(user_id, result_rel)
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(result_bytes))
        ws = wb[cfg["sheet_name"]] if cfg.get("sheet_name") else wb.active

        for row in range(cfg["start_row"], cfg["end_row"] + 1):
            if task["_cancelled"]:
                task["status"] = "cancelled"
                task["finished_at"] = datetime.now().isoformat()
                buf = io.BytesIO()
                wb.save(buf)
                storage.write_bytes(user_id, result_rel, buf.getvalue())
                return

            query = ws.cell(row=row, column=query_col_idx).value
            if not query or not str(query).strip():
                task["results"].append({"row": row, "query": "", "content": "", "tool_calls": "", "status": "skipped"})
                task["completed"] += 1
                continue

            query = str(query).strip()
            task["current_query"] = query[:100]

            try:
                thread_id = f"batch-{task_id}-row{row}"
                result = await agent.ainvoke(
                    {"messages": [{"role": "user", "content": query}]},
                    config={"configurable": {"thread_id": thread_id}},
                )
                content_text = ""
                tool_calls_text = ""
                for msg in reversed(result.get("messages", [])):
                    if msg.__class__.__name__ == "AIMessage" and not content_text:
                        c = msg.content
                        if isinstance(c, str):
                            content_text = c
                        elif isinstance(c, list):
                            content_text = " ".join(b.get("text", "") for b in c if isinstance(b, dict) and b.get("type") == "text")
                        tc = getattr(msg, "tool_calls", None) or []
                        if tc:
                            tool_calls_text = "; ".join(f"{t.get('name', '')}({json.dumps(t.get('args', {}), ensure_ascii=False)[:120]})" for t in tc)
                        break

                all_tool_calls = []
                for msg in result.get("messages", []):
                    if msg.__class__.__name__ == "AIMessage":
                        for t in (getattr(msg, "tool_calls", None) or []):
                            all_tool_calls.append(f"{t.get('name', '')}({json.dumps(t.get('args', {}), ensure_ascii=False)[:120]})")
                if all_tool_calls:
                    tool_calls_text = "; ".join(all_tool_calls)

                ws.cell(row=row, column=content_col_idx, value=content_text[:32000])
                ws.cell(row=row, column=tool_col_idx, value=tool_calls_text[:32000])
                task["results"].append({"row": row, "query": query[:100], "content": content_text[:200], "tool_calls": tool_calls_text[:200], "status": "done"})
            except Exception as e:
                error_msg = str(e)[:200]
                ws.cell(row=row, column=content_col_idx, value=f"[ERROR] {error_msg}")
                task["results"].append({"row": row, "query": query[:100], "content": f"[ERROR] {error_msg}", "tool_calls": "", "status": "error"})

            task["completed"] += 1
            if task["completed"] % 5 == 0:
                buf = io.BytesIO()
                wb.save(buf)
                storage.write_bytes(user_id, result_rel, buf.getvalue())

        buf = io.BytesIO()
        wb.save(buf)
        storage.write_bytes(user_id, result_rel, buf.getvalue())
        task["status"] = "completed"
        task["finished_at"] = datetime.now().isoformat()
        task["current_query"] = ""
    except Exception as e:
        task["status"] = "error"
        task["error"] = str(e)[:500]
        task["finished_at"] = datetime.now().isoformat()
    finally:
        _start_next_queued_task(user_id)


def _start_next_queued_task(user_id: str):
    for t in _batch_tasks.values():
        if t["user_id"] == user_id and t["status"] == "queued" and not t["_cancelled"]:
            t["status"] = "running"
            asyncio.create_task(_run_batch_task(t["id"], user_id, t["config"]["model"], t["_prompt_content"]))
            return
